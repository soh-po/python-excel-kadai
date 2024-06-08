import openpyxl
import pandas as pd

# ファイルの読み込みとデータの連結
file_list = ["2022_年間売上表.xlsx", "2023_年間売上表.xlsx"]

df = pd.DataFrame()

for file_name in file_list:
    temp_df = pd.read_excel(file_name, sheet_name="Sheet1", header=None, skiprows=1)
    temp_df.columns = ["売上年", "商品", "金額（千円）"]
    df = pd.concat([df, temp_df], ignore_index=True)

# データの集約
grouped = df.groupby(["商品", "売上年"]).agg({"金額（千円）":"sum"}).reset_index()

# pandasでファイルに保存
file_name = "売上集計表.xlsx"
writer = pd.ExcelWriter(file_name)
grouped.to_excel(writer, index=False)
writer.close()

# openypxlで開く
wb = openpyxl.load_workbook(file_name)
ws = wb.active

# 書式設定
color = openpyxl.styles.PatternFill(patternType="solid", fgColor="F2F2F2")

ws["A1"].fill = color
ws["B1"].fill = color
ws["C1"].fill = color

# 保存
wb.save(file_name)
