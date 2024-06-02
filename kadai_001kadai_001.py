import openpyxl
import time


title = "請求書"
company = "株式会社ABC"
address = "〒101-0022 東京都千代田区神田練塀町300"
phone = "TEL:03-1234-5678 FAX:03-1234-5678"
name = "担当者名:鈴木一郎 様"
today = time.strftime("%Y/%m/%d")

filename = f"請求書_{time.strftime("%Y%m%d")}.xlsx"


header = ["商品名", "数量", "単価", "金額"]
data = [
    ["商品A", 2, 10000, "=C11*D11"],
    ["商品B", 1, 15000, "=C12*D12"],
    ["", "", "", "=SUM(E11:E12)"]
]
tax_calc = [
    ["小計", "", "", "=SUM(E11:E12)"],
    ["消費税", "", "", "=E13*0.1"],
    ["合計", "", "", "=SUM(E15:E16)"]
]

wb = openpyxl.Workbook()
ws = wb.active

ws["A2"] = title
ws["A4"] = company
ws["E4"] = "No."
ws["F4"] = "0001"
ws["E5"] = "日付"
ws["F5"] = today
ws["A5"] = address
ws["A6"] = phone
ws["A7"] = name


# heaerの挿入
ws.append(header)
ws.insert_rows(8, 2)

# 製品情報の記入
for row in data:
    ws.append(row)


for row in tax_calc:
    ws.append(row)


ws.insert_cols(1)
ws.insert_rows(14, 1)

wb.save(filename)
