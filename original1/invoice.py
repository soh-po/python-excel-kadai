from glob import glob
from datetime import datetime
import openpyxl as opx
import pandas as pd


filename_sales = glob("salesbooks/*.xlsx") # フォルダ内のxlsxファイルをリストに入れる
template_file = "invoice-template.xlsx" # 請求書作成に使うテンプレートファイル
# today = datetime.now() # 日時（当日）を取得
# cell_date = today.strftime("%Y年%m月%d日") # プログラム実行時に日付にする場合に使う
cell_date = "2024年4月30日" # 任意の日付を指定

df = pd.DataFrame(columns=["日付","購入者","品目", "個数", "値段", "小計"]) # データフレームの列を定義


# openpyxlでファイルを読み込み、pandasでdfにまとめる
row = 0
for files in filename_sales: # xlsxファイルを順に変数filesに代入
    # print(files)
    wb_sales = opx.load_workbook(files, data_only=True) # xlsxファイルを数値で読み込む（Excelの計算式から計算結果を数値で書き込む）
    ws_sales = wb_sales["3月"] # 所定のワークシートを選択
    ws_title = ws_sales.title # シート名の取得（請求月表示に使用）
    max_row = ws_sales.max_row # シートの最終行を取得  

    for r in range(4, max_row + 1):
        if ws_sales.cell(r, 1).value is not None:
            df.loc[row, "日付"] = ws_sales.cell(r, 1).value
            df.loc[row, "購入者"] = ws_sales.cell(r, 2).value
            df.loc[row, "品目"] = ws_sales.cell(r, 3).value
            df.loc[row, "個数"] = ws_sales.cell(r, 4).value
            df.loc[row, "値段"] = ws_sales.cell(r, 5).value
            df.loc[row, "小計"] = ws_sales.cell(r, 6).value
            row += 1

grouped = df.groupby("購入者") # 購入者毎にグループ化

customers = {}
for name, group in grouped:
    customers[name] = group.reset_index(drop=True)

for name, customer_df in customers.items():
    wb = opx.load_workbook(template_file)
    ws = wb.active

    for i, row in customer_df.iterrows():
        ws["B4"] = row["購入者"]
        file_name = row["購入者"]
        # without_space_file_name = file_name.replace(" ", "") # ファイル名からスペースを削除する場合
        ws["G3"] = cell_date # 請求書の日付
        ws["C10"] = f"{ws_title}分のご請求" # 件名
        bond_value = f"{row['品目']}({row['日付'].strftime('%m/%d')})" # 品目と日付を1つのセルに書く(内訳欄)変数を定義
        ws.cell(row=i + 15, column=2, value=bond_value) # 内訳欄
        ws.cell(row=i + 15, column=5, value=row["個数"]) # 個数欄
        ws.cell(row=i + 15, column=6, value=row["値段"]) # 単価欄
        ws.cell(row=i + 15, column=7, value=row["小計"]) # 金額(税込)欄
        # カーソルの位置などを調整する場合
        # ws.sheet_view.selection[0].activeCell = "A1"
        # ws.sheet_view.selection[0].sqref = "A1"
        # ws.sheet_view.topLeftCell = "A1"
        wb.save(f"invoice/{file_name}様.xlsx") # ファイルに保存

print("請求書を作成しました。")
